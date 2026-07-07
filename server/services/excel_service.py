"""Excel import/export service."""

import json
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.models import (
    AnnualEvaluation,
    EvaluationRule,
    Organization,
    QuarterlyEvaluation,
)


def _style_header(ws, row: int, max_col: int) -> None:
    """Apply header styling."""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


async def export_quarterly_excel(
    session: AsyncSession, year: int, quarter: int
) -> BytesIO:
    """Export quarterly evaluation data to an Excel file matching the template format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{quarter}季度评价结果"

    # Title
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1, value=f"{year}年第{quarter}季度基层党组织创先争优评价结果")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "序号", "分组", "党组织名称",
        "「七大动力」评价得分", "「七大动力」扣分", "「七大动力」扣分说明",
        "季度党群重点工作得分", "季度党群重点工作扣分", "季度党群重点工作扣分说明",
        "表彰奖励", "重点课题", "经验交流", "品牌建设",
        "新闻宣传", "巡察纪委", "工会贡献", "团组织贡献",
        "贡献度合计", "党群绩效总分",
        "党群绩效结果", "行政绩效结果", "创先争优结果",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(headers))

    # Data
    result = await session.execute(
        select(QuarterlyEvaluation)
        .where(
            QuarterlyEvaluation.year == year,
            QuarterlyEvaluation.quarter == quarter,
        )
        .order_by(QuarterlyEvaluation.id)
    )
    evals: list[QuarterlyEvaluation] = list(result.scalars().all())

    for i, ev in enumerate(evals, 1):
        org = await session.get(Organization, ev.org_id)
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=org.group if org else "")
        ws.cell(row=row, column=3, value=org.name if org else "")
        ws.cell(row=row, column=4, value=ev.score_7powers_base - ev.deduction_7powers)
        ws.cell(row=row, column=5, value=ev.deduction_7powers)
        ws.cell(row=row, column=6, value=ev.deduction_reason_7powers)
        ws.cell(row=row, column=7, value=ev.score_key_work_base - ev.deduction_key_work)
        ws.cell(row=row, column=8, value=ev.deduction_key_work)
        ws.cell(row=row, column=9, value=ev.deduction_reason_key_work)
        ws.cell(row=row, column=10, value=ev.score_commend)
        ws.cell(row=row, column=11, value=ev.score_topic)
        ws.cell(row=row, column=12, value=ev.score_exchange)
        ws.cell(row=row, column=13, value=ev.score_brand)
        ws.cell(row=row, column=14, value=ev.score_media)
        ws.cell(row=row, column=15, value=ev.score_inspect)
        ws.cell(row=row, column=16, value=ev.score_union)
        ws.cell(row=row, column=17, value=ev.score_youth)
        ws.cell(row=row, column=18, value=ev.contribution_total)
        ws.cell(row=row, column=19, value=round(ev.total_score, 2))
        ws.cell(row=row, column=20, value=ev.party_mass_grade)
        ws.cell(row=row, column=21, value=ev.admin_grade)
        ws.cell(row=row, column=22, value=ev.final_grade)

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_annual_excel(session: AsyncSession, year: int) -> BytesIO:
    """Export annual evaluation data to an Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年度评价结果"

    ws.merge_cells("A1:O1")
    title_cell = ws.cell(row=1, column=1, value=f"{year}年度基层党组织创先争优评价结果")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    headers = [
        "序号", "分组", "党组织名称",
        "一季度得分", "二季度得分", "三季度得分", "四季度得分",
        "季度平均分", "季度权重分(60%)",
        "年度工作得分", "年度工作权重分(30%)",
        "满意度得分", "满意度权重分(10%)",
        "年度总分",
        "党群绩效结果", "行政绩效结果", "创先争优结果",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(headers))

    result = await session.execute(
        select(AnnualEvaluation)
        .where(AnnualEvaluation.year == year)
        .order_by(AnnualEvaluation.id)
    )
    evals: list[AnnualEvaluation] = list(result.scalars().all())

    for i, ev in enumerate(evals, 1):
        org = await session.get(Organization, ev.org_id)
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=org.group if org else "")
        ws.cell(row=row, column=3, value=org.name if org else "")
        ws.cell(row=row, column=4, value=round(ev.q1_total_score, 2))
        ws.cell(row=row, column=5, value=round(ev.q2_total_score, 2))
        ws.cell(row=row, column=6, value=round(ev.q3_total_score, 2))
        ws.cell(row=row, column=7, value=round(ev.q4_total_score, 2))
        ws.cell(row=row, column=8, value=round(ev.quarterly_avg, 2))
        ws.cell(row=row, column=9, value=round(ev.quarterly_weighted, 4))
        ws.cell(row=row, column=10, value=round(ev.annual_work_score, 2))
        ws.cell(row=row, column=11, value=round(ev.annual_work_weighted, 4))
        ws.cell(row=row, column=12, value=round(ev.satisfaction_score, 2))
        ws.cell(row=row, column=13, value=round(ev.satisfaction_weighted, 4))
        ws.cell(row=row, column=14, value=round(ev.total_score, 4))
        ws.cell(row=row, column=15, value=ev.party_mass_grade)
        ws.cell(row=row, column=16, value=ev.admin_grade)
        ws.cell(row=row, column=17, value=ev.final_grade)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
